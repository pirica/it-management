import openpyxl
from openpyxl.styles import Font, PatternFill

def create_asset_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"

    headers = ["name", "brand", "model", "serial_number", "category", "purchase_date", "purchase_cost"]
    required = ["name"]

    ws.append(headers)

    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    required_font = Font(bold=True, color="FF0000")

    for cell in ws[1]:
        cell.fill = header_fill
        if cell.value in required:
            cell.font = required_font

    # Add a sample row
    ws.append(["Dell Precision 5550", "Dell Technologies", "5550", "SN-5550-XYZ", "Workstation", "2023-01-15", "1500.00"])

    wb.save("modules/import/asset_template.xlsx")

def create_employee_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = ["full_name", "email", "username", "phone", "job_title", "employee_id"]
    required = ["full_name", "email"]

    ws.append(headers)

    header_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    required_font = Font(bold=True, color="FF0000")

    for cell in ws[1]:
        cell.fill = header_fill
        if cell.value in required:
            cell.font = required_font

    # Add a sample row
    ws.append(["John Smith", "john@example.com", "jsmith", "+123456789", "Software Engineer", "EMP001"])

    wb.save("modules/import/employee_template.xlsx")

if __name__ == "__main__":
    create_asset_template()
    create_employee_template()
